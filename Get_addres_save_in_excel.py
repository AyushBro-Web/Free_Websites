import os
from flask import Flask, request
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Define the Excel file path
excel_file_path = '/storage/emulated/0/Address.xlsx'

# Create the Excel file with headers if it doesn't exist
if not os.path.exists(excel_file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Address', 'City', 'State', 'Pincode', 'Email', 'Ph', 'Alternative Ph'])
    wb.save(excel_file_path)
else:
    wb = load_workbook(excel_file_path)
    ws = wb.active
    if ws.max_row == 1 and all(ws.cell(row=1, column=i).value is None for i in range(1, 9)):
        ws.append(['Name', 'Address', 'City', 'State', 'Pincode', 'Email', 'Ph', 'Alternative Ph'])
    wb.save(excel_file_path)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        data = request.form
        name = data['name']
        address = data['address']
        city = data['city']
        state = data['state']
        pincode = data['pincode']
        email = data['email']
        phone = data['ph']
        alternative_phone = data['alternative_ph']

        # Log user information to console (without password)
        print(f"Received form submission - Name: {name}, Address: {address}, City: {city}, State: {state}, Pincode: {pincode}, Email: {email}, Ph: {phone}, Alternative Ph: {alternative_phone}")

        # Save the form data into the Excel file
        wb = load_workbook(excel_file_path)
        ws = wb.active
        ws.append([name, address, city, state, pincode, email, phone, alternative_phone])
        wb.save(excel_file_path)

        # Display a thank you message after submission
        return '''
        <div style="background-color:green;color:aliceblue; padding:20px;text-align:center;">
            <h1>Thanks for Response!</h1>
            <h4>We have received your information and will contact you soon.</h4>
        </div>
        <div style="background-color:skyblue;color:green; padding:20px;text-align:center;">
            <h3>For other queries, contact us tomorrow. Thank you. YOUR BROTHER AYUSH</h3>
        </div>
        <div class="footer">
            <marquee><h3>For more information | Contact us at: Your Bussiness Accounts and Gmails| Thank you for choosing our service!</h3></marquee>
        </div>
        ''', 200

    return '''
    <!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Free Website.in</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 0;
            background-color: yellow;
            color: black;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            height: 100vh;
            overflow: hidden;
        }
        .container {
            max-width: 800px;
            width: 100%;
            text-align: center;
            flex-grow: 1;
        }
        .form-container {
            text-align: left;
            display: inline-block;
            background-color: #fff;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
        }
        .form-container label {
            display: block;
            margin: 10px 0 5px;
        }
        .form-container input, .form-container textarea {
            width: calc(100% - 22px);
            padding: 10px;
            margin: 5px 0;
            border: 1px solid #ddd;
            border-radius: 4px;
        }
        .form-container button {
            padding: 12px 24px;
            font-size: 16px;
            color: #fff;
            background-color: #007BFF;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            transition: background-color 0.3s ease;
        }
        .form-container button:hover {
            background-color: #0056b3;
        }

        .footer {
            width: 100%;
            background-color: #333;
            color: white;
            text-align: center;
            position: fixed;
            bottom: 0;
            left: 0;
            padding: 10px 0;
            overflow: hidden;
        }

        .footer marquee {
            font-size: 1em;
            white-space: nowrap;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Contact Us</h1>
        <div class="form-container">
            <form method="post">
                <label for="name">Name:</label>
                <input type="text" id="name" name="name" required>
                
                <label for="address">Address:</label>
                <input type="text" id="address" name="address" required>

                <label for="city">City:</label>
                <input type="text" id="city" name="city" required>

                <label for="state">State:</label>
                <input type="text" id="state" name="state" required>

                <label for="pincode">Pincode:</label>
                <input type="number" id="pincode" name="pincode" required>

                <label for="email">Email:</label>
                <input type="email" id="email" name="email" required>

                <label for="ph">Ph:</label>
                <input type="number" id="ph" name="ph" required>

                <label for="alternative_ph">Alternative Ph:</label>
                <input type="number" id="alternative_ph" name="alternative_ph">

                <button type="submit" align='center'>Submit</button>
            </form>
        </div>
    </div>
    <div class="footer">
        <marquee>© 2024 The Web Developers!</marquee>
    </div>
</body>
</html>
    '''

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000)
